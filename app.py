import os
import re
import sqlite3
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# Excel 导出：需要 openpyxl
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


APP_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(APP_DIR, "data.db")
ATTACH_DIR = os.path.join(APP_DIR, "attachments")
os.makedirs(ATTACH_DIR, exist_ok=True)


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def norm_imei(s: str) -> str:
    """Normalize IMEI: keep digits only."""
    if not s:
        return ""
    digits = re.sub(r"\D+", "", s.strip())
    return digits


class DB:
    def __init__(self, path: str):
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.init()

    def init(self):
        cur = self.conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            date TEXT NOT NULL,
            brand TEXT,
            model TEXT,
            imei1 TEXT,
            imei2 TEXT,
            color TEXT,
            storage TEXT,
            condition TEXT,
            issues TEXT,
            buy_price REAL,
            payment TEXT,
            seller_name TEXT,
            seller_doc TEXT,
            seller_phone TEXT,
            notes TEXT,
            attachments TEXT
        );
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_records_imei1 ON records(imei1);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_records_imei2 ON records(imei2);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_records_date ON records(date);")
        self.conn.commit()

    def imei_exists(self, imei: str, exclude_id=None) -> bool:
        if not imei:
            return False
        cur = self.conn.cursor()
        if exclude_id is None:
            cur.execute("SELECT 1 FROM records WHERE imei1=? OR imei2=? LIMIT 1", (imei, imei))
        else:
            cur.execute(
                "SELECT 1 FROM records WHERE (imei1=? OR imei2=?) AND id<>? LIMIT 1",
                (imei, imei, exclude_id)
            )
        return cur.fetchone() is not None

    def insert(self, data: dict) -> int:
        keys = list(data.keys())
        vals = [data[k] for k in keys]
        placeholders = ",".join(["?"] * len(keys))
        sql = f"INSERT INTO records ({','.join(keys)}) VALUES ({placeholders})"
        cur = self.conn.cursor()
        cur.execute(sql, vals)
        self.conn.commit()
        return cur.lastrowid

    def update(self, rec_id: int, data: dict):
        keys = list(data.keys())
        vals = [data[k] for k in keys]
        set_clause = ", ".join([f"{k}=?" for k in keys])
        sql = f"UPDATE records SET {set_clause} WHERE id=?"
        cur = self.conn.cursor()
        cur.execute(sql, vals + [rec_id])
        self.conn.commit()

    def delete(self, rec_id: int):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM records WHERE id=?", (rec_id,))
        self.conn.commit()

    def get(self, rec_id: int):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM records WHERE id=?", (rec_id,))
        return cur.fetchone()

    def search(self, kw: str = "", date_from: str = "", date_to: str = ""):
        kw = (kw or "").strip()
        params = []
        where = []

        if kw:
            where.append(
                "(brand LIKE ? OR model LIKE ? OR imei1 LIKE ? OR imei2 LIKE ? "
                "OR seller_name LIKE ? OR seller_phone LIKE ? OR notes LIKE ?)"
            )
            like = f"%{kw}%"
            params.extend([like, like, like, like, like, like, like])

        if date_from:
            where.append("date >= ?")
            params.append(date_from.strip())

        if date_to:
            where.append("date <= ?")
            params.append(date_to.strip())

        sql = "SELECT * FROM records"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY id DESC LIMIT 5000"

        cur = self.conn.cursor()
        cur.execute(sql, params)
        return cur.fetchall()

    def all_for_export(self):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM records ORDER BY id DESC")
        return cur.fetchall()


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        # ===== 方案1：默认更大 + 启动最大化 =====
        self.title("二手机回收记录（离线版）")
        self.geometry("1400x820")
        self.minsize(1200, 700)
        try:
            self.state("zoomed")  # Win10/11 最大化
        except Exception:
            pass
        # ====================================

        self.db = DB(DB_PATH)
        self.selected_id = None
        self.attachment_paths = []
        self.current_attachments_saved = ""

        self._build_ui()
        self.refresh_list()

    def _build_ui(self):
        # Top: Search
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Label(top, text="搜索（IMEI/品牌/型号/姓名/电话/备注）:").pack(side="left")
        self.search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.search_var, width=35).pack(side="left", padx=6)

        ttk.Label(top, text="日期从(YYYY-MM-DD):").pack(side="left", padx=(12, 0))
        self.date_from_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.date_from_var, width=12).pack(side="left", padx=6)

        ttk.Label(top, text="到:").pack(side="left")
        self.date_to_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.date_to_var, width=12).pack(side="left", padx=6)

        ttk.Button(top, text="查询", command=self.refresh_list).pack(side="left", padx=6)
        ttk.Button(top, text="清空条件", command=self.clear_search).pack(side="left", padx=6)

        ttk.Button(top, text="导出Excel", command=self.export_excel).pack(side="right")
        ttk.Button(top, text="打开附件文件夹", command=self.open_attach_dir).pack(side="right", padx=8)

        # Middle: Left list + Right form
        mid = ttk.Frame(self)
        mid.pack(fill="both", expand=True, padx=10, pady=8)

        left = ttk.Frame(mid)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        # ===== 方案2：右侧固定宽度，禁止被挤压 =====
        right = ttk.Frame(mid, width=360)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)
        # ===========================================

        # Treeview
        cols = ("id", "date", "brand", "model", "imei1", "buy_price", "payment", "seller_name", "seller_phone")
        self.tree = ttk.Treeview(left, columns=cols, show="headings", height=18)
        headings = {
            "id": "ID",
            "date": "日期",
            "brand": "品牌",
            "model": "型号",
            "imei1": "IMEI1",
            "buy_price": "回收价€",
            "payment": "支付",
            "seller_name": "卖家姓名",
            "seller_phone": "电话",
        }
        widths = {
            "id": 60,
            "date": 120,
            "brand": 100,
            "model": 140,
            "imei1": 180,
            "buy_price": 100,
            "payment": 80,
            "seller_name": 120,
            "seller_phone": 120
        }
        for c in cols:
            self.tree.heading(c, text=headings[c])
            self.tree.column(c, width=widths[c], anchor="w")

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

        yscroll = ttk.Scrollbar(left, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=yscroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        yscroll.pack(side="right", fill="y")

        # Form
        form = ttk.LabelFrame(right, text="记录详情（新增/编辑）")
        form.pack(fill="both", expand=True, padx=6, pady=6)

        self.vars = {}

        def add_row(label, key, width=24):
            row = ttk.Frame(form)
            row.pack(fill="x", pady=3, padx=8)
            ttk.Label(row, text=label, width=10).pack(side="left")
            v = tk.StringVar()
            ent = ttk.Entry(row, textvariable=v, width=width)
            ent.pack(side="left", fill="x", expand=True)
            self.vars[key] = v
            return ent

        # Default date today
        self.vars["date"] = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        row = ttk.Frame(form)
        row.pack(fill="x", pady=3, padx=8)
        ttk.Label(row, text="日期", width=10).pack(side="left")
        ttk.Entry(row, textvariable=self.vars["date"], width=24).pack(side="left", fill="x", expand=True)

        add_row("品牌", "brand")
        add_row("型号", "model")
        add_row("IMEI1", "imei1")
        add_row("IMEI2", "imei2")
        add_row("颜色", "color")
        add_row("容量", "storage")
        add_row("成色/等级", "condition")
        add_row("故障/问题", "issues")

        # Price
        rowp = ttk.Frame(form)
        rowp.pack(fill="x", pady=3, padx=8)
        ttk.Label(rowp, text="回收价€", width=10).pack(side="left")
        self.vars["buy_price"] = tk.StringVar()
        ttk.Entry(rowp, textvariable=self.vars["buy_price"], width=24).pack(side="left", fill="x", expand=True)

        # Payment dropdown
        rowpay = ttk.Frame(form)
        rowpay.pack(fill="x", pady=3, padx=8)
        ttk.Label(rowpay, text="支付方式", width=10).pack(side="left")
        self.vars["payment"] = tk.StringVar(value="现金")
        ttk.Combobox(
            rowpay,
            textvariable=self.vars["payment"],
            values=["现金", "转账", "POS"],
            width=22,
            state="readonly"
        ).pack(side="left", fill="x", expand=True)

        add_row("卖家姓名", "seller_name")
        add_row("证件号", "seller_doc")
        add_row("电话", "seller_phone")

        # Notes
        rown = ttk.Frame(form)
        rown.pack(fill="both", pady=3, padx=8)
        ttk.Label(rown, text="备注", width=10).pack(side="left", anchor="n")
        self.notes_text = tk.Text(rown, width=24, height=5)
        self.notes_text.pack(side="left", fill="both", expand=True)

        # Attachments
        rowa = ttk.Frame(form)
        rowa.pack(fill="x", pady=6, padx=8)
        ttk.Label(rowa, text="附件", width=10).pack(side="left")
        ttk.Button(rowa, text="添加照片/文件", command=self.add_attachments).pack(side="left")
        ttk.Button(rowa, text="查看附件", command=self.view_attachments).pack(side="left", padx=6)

        self.attach_label = ttk.Label(form, text="(未添加)", wraplength=320, foreground="#555")
        self.attach_label.pack(fill="x", padx=10, pady=(0, 6))

        # Buttons
        btns = ttk.Frame(form)
        btns.pack(fill="x", padx=8, pady=8)

        ttk.Button(btns, text="新建/清空", command=self.new_record).pack(side="left")
        ttk.Button(btns, text="保存", command=self.save_record).pack(side="left", padx=6)
        ttk.Button(btns, text="删除", command=self.delete_record).pack(side="left", padx=6)

        tip = ttk.Label(form, text="提示：IMEI 会自动去掉空格/符号并查重。", foreground="#777")
        tip.pack(fill="x", padx=10, pady=(0, 8))

    def clear_search(self):
        self.search_var.set("")
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh_list()

    def refresh_list(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        rows = self.db.search(
            kw=self.search_var.get(),
            date_from=self.date_from_var.get(),
            date_to=self.date_to_var.get()
        )
        for r in rows:
            self.tree.insert("", "end", values=(
                r["id"],
                r["date"],
                r["brand"] or "",
                r["model"] or "",
                r["imei1"] or "",
                f'{r["buy_price"]:.2f}' if r["buy_price"] is not None else "",
                r["payment"] or "",
                r["seller_name"] or "",
                r["seller_phone"] or ""
            ))

    def on_select(self, _evt=None):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        rec_id = int(vals[0])
        rec = self.db.get(rec_id)
        if not rec:
            return

        self.selected_id = rec_id
        self.vars["date"].set(rec["date"] or "")
        for k in ["brand", "model", "imei1", "imei2", "color", "storage", "condition", "issues",
                  "buy_price", "payment", "seller_name", "seller_doc", "seller_phone"]:
            if k == "buy_price":
                self.vars[k].set("" if rec[k] is None else str(rec[k]))
            else:
                self.vars[k].set(rec[k] or "")

        self.notes_text.delete("1.0", "end")
        self.notes_text.insert("1.0", rec["notes"] or "")

        self.attachment_paths = []
        self.current_attachments_saved = rec["attachments"] or ""
        self._update_attach_label()

    def _update_attach_label(self):
        parts = []
        if self.current_attachments_saved:
            parts.extend([p for p in self.current_attachments_saved.split(";") if p.strip()])
        if self.attachment_paths:
            parts.extend([os.path.basename(p) for p in self.attachment_paths])
        self.attach_label.config(text="(未添加)" if not parts else "; ".join(parts))

    def new_record(self):
        self.selected_id = None
        self.vars["date"].set(datetime.now().strftime("%Y-%m-%d"))
        for k in ["brand", "model", "imei1", "imei2", "color", "storage", "condition", "issues",
                  "buy_price", "seller_name", "seller_doc", "seller_phone"]:
            self.vars[k].set("")
        self.vars["payment"].set("现金")
        self.notes_text.delete("1.0", "end")
        self.attachment_paths = []
        self.current_attachments_saved = ""
        self._update_attach_label()

    def add_attachments(self):
        paths = filedialog.askopenfilenames(
            title="选择附件（身份证照片/手机照片/收据等）",
            filetypes=[("所有文件", "*.*")]
        )
        if paths:
            self.attachment_paths.extend(paths)
            self._update_attach_label()

    def _copy_attachments(self, rec_id: int) -> str:
        saved = []
        if self.current_attachments_saved:
            saved.extend([p for p in self.current_attachments_saved.split(";") if p.strip()])

        if not self.attachment_paths:
            return ";".join(saved)

        rec_folder = os.path.join(ATTACH_DIR, f"ID_{rec_id}")
        os.makedirs(rec_folder, exist_ok=True)

        for src in self.attachment_paths:
            base = os.path.basename(src)
            dst = os.path.join(rec_folder, base)
            if os.path.exists(dst):
                name, ext = os.path.splitext(base)
                dst = os.path.join(rec_folder, f"{name}_{int(datetime.now().timestamp())}{ext}")
            try:
                with open(src, "rb") as fsrc, open(dst, "wb") as fdst:
                    fdst.write(fsrc.read())
                rel = os.path.relpath(dst, APP_DIR)
                saved.append(rel)
            except Exception as e:
                messagebox.showwarning("附件复制失败", f"无法复制附件：{src}\n原因：{e}")

        self.attachment_paths = []
        return ";".join(saved)

    def view_attachments(self):
        if self.selected_id is None:
            messagebox.showinfo("提示", "请先选择一条记录（或保存后再查看附件）。")
            return
        rec = self.db.get(self.selected_id)
        if not rec:
            return
        att = rec["attachments"] or ""
        if not att.strip():
            messagebox.showinfo("附件", "这条记录没有附件。")
            return

        folder = os.path.join(ATTACH_DIR, f"ID_{self.selected_id}")
        if os.path.isdir(folder):
            os.startfile(folder)
        else:
            messagebox.showinfo("附件", "附件文件夹不存在（可能附件路径已被移动）。")

    def open_attach_dir(self):
        os.startfile(ATTACH_DIR)

    def save_record(self):
        data = {
            "date": (self.vars["date"].get() or "").strip(),
            "brand": (self.vars["brand"].get() or "").strip(),
            "model": (self.vars["model"].get() or "").strip(),
            "imei1": norm_imei(self.vars["imei1"].get()),
            "imei2": norm_imei(self.vars["imei2"].get()),
            "color": (self.vars["color"].get() or "").strip(),
            "storage": (self.vars["storage"].get() or "").strip(),
            "condition": (self.vars["condition"].get() or "").strip(),
            "issues": (self.vars["issues"].get() or "").strip(),
            "payment": (self.vars["payment"].get() or "").strip(),
            "seller_name": (self.vars["seller_name"].get() or "").strip(),
            "seller_doc": (self.vars["seller_doc"].get() or "").strip(),
            "seller_phone": (self.vars["seller_phone"].get() or "").strip(),
            "notes": self.notes_text.get("1.0", "end").strip(),
        }

        if not data["date"]:
            messagebox.showerror("错误", "日期不能为空（例如 2026-02-10）。")
            return

        price_raw = (self.vars["buy_price"].get() or "").strip().replace(",", ".")
        if price_raw:
            try:
                data["buy_price"] = float(price_raw)
            except ValueError:
                messagebox.showerror("错误", "回收价格式不对，请输入数字，例如 120 或 120.5")
                return
        else:
            data["buy_price"] = None

        for imei in [data["imei1"], data["imei2"]]:
            if imei and len(imei) < 10:
                messagebox.showwarning("提示", f"IMEI 看起来太短：{imei}\n你可以继续保存，但建议核对。")

        if data["imei1"] and self.db.imei_exists(data["imei1"], exclude_id=self.selected_id):
            if not messagebox.askyesno("IMEI重复", f"IMEI1（{data['imei1']}）已存在记录！\n仍然要保存吗？"):
                return
        if data["imei2"] and self.db.imei_exists(data["imei2"], exclude_id=self.selected_id):
            if not messagebox.askyesno("IMEI重复", f"IMEI2（{data['imei2']}）已存在记录！\n仍然要保存吗？"):
                return

        if self.selected_id is None:
            data_full = {"created_at": now_str(), "attachments": "", **data}
            rec_id = self.db.insert(data_full)
            att = self._copy_attachments(rec_id)
            self.db.update(rec_id, {"attachments": att})
            self.selected_id = rec_id
            self.current_attachments_saved = att
            self._update_attach_label()
            messagebox.showinfo("成功", f"已新增记录 ID={rec_id}")
        else:
            att = self._copy_attachments(self.selected_id)
            data["attachments"] = att
            self.db.update(self.selected_id, data)
            self.current_attachments_saved = att
            self._update_attach_label()
            messagebox.showinfo("成功", f"已保存记录 ID={self.selected_id}")

        self.refresh_list()

    def delete_record(self):
        if self.selected_id is None:
            messagebox.showinfo("提示", "请先选择要删除的记录。")
            return
        if not messagebox.askyesno("确认删除", f"确定删除 ID={self.selected_id} 这条记录吗？\n（数据库记录会删除，附件文件不会自动删）"):
            return
        self.db.delete(self.selected_id)
        messagebox.showinfo("已删除", f"已删除 ID={self.selected_id}")
        self.new_record()
        self.refresh_list()

    def export_excel(self):
        if openpyxl is None:
            messagebox.showerror("缺少组件", "导出Excel需要安装 openpyxl。\n请先运行：pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            title="保存为 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not path:
            return

        rows = self.db.all_for_export()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "回收记录"

        headers = [
            "ID", "创建时间", "日期", "品牌", "型号", "IMEI1", "IMEI2", "颜色", "容量", "成色/等级", "故障/问题",
            "回收价€", "支付方式", "卖家姓名", "证件号", "电话", "备注", "附件(相对路径)"
        ]
        ws.append(headers)

        for r in rows:
            ws.append([
                r["id"], r["created_at"], r["date"], r["brand"], r["model"], r["imei1"], r["imei2"],
                r["color"], r["storage"], r["condition"], r["issues"],
                r["buy_price"], r["payment"], r["seller_name"], r["seller_doc"], r["seller_phone"],
                r["notes"], r["attachments"]
            ])

        for col in range(1, len(headers) + 1):
            maxlen = 10
            for cell in ws[get_column_letter(col)]:
                if cell.value is None:
                    continue
                maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(maxlen + 2, 45)

        try:
            wb.save(path)
            messagebox.showinfo("导出成功", f"已导出：\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", f"无法保存Excel：{e}")


if __name__ == "__main__":
    # nicer on Windows
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    app = App()
    app.mainloop()
